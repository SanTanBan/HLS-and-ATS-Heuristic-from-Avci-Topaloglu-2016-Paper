import time
import matplotlib.pylab as plt
import pulp as p
#import winsound
import os
import openpyxl
import GeneRead

def PuLP_Algorithm(max_seconds_allowed_for_calculation=0,directory_containing_Node_Specifications_file="",directory__containing_Vehicle_Specifications_file="",directory_containing_Distance_Matrices_file="",directory_details_for_saving=""):
    
    os.mkdir(directory_details_for_saving+"PuLP Solution")
    directory_to_save_PuLP_solution=directory_details_for_saving+"PuLP Solution/"

    Depot_and_Relief_Centres,Latitude,Longitude,PickUp,Delivery=GeneRead.Reader.Relief_Centre_Reader(directory_of_Relief_Centre_Specifications_file=directory_containing_Node_Specifications_file)
    C=GeneRead.Reader.Distance_Combinations_Reader(directory_containing_Distance_Locations_file=directory_containing_Distance_Matrices_file,directory_of_Vehicle_Types_considered_file=directory__containing_Vehicle_Specifications_file)
    Vehicle_Types,VN,VQ,VS,VC,Vehicle_Width=GeneRead.Reader.Vehicle_Type_Reader(directory_of_Vehicle_Type_file=directory__containing_Vehicle_Specifications_file)
    Relief_Centres=Depot_and_Relief_Centres.copy()
    del Relief_Centres[0]
    Vehicle_Type_Maximum_Utilised_Capacity={}
    Number_of_Vehicle_used_of_each_Type={}


    #Solving the problem using PuLP within the Time Limit
    # Set the problem
    prob=p.LpProblem("Heterogenenous_single_Depot_mVRPSDP",p.LpMinimize)

    # Decision Variables
    # Iff Arc joining i & j is included within the solution for the Layer k
    x=p.LpVariable.dicts('x',((i,j,k) for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres for k in Vehicle_Types if i!=j),cat='Binary')

    # Amount of collected load across Arc(i,j) by a Vehicle in Layer k
    y=p.LpVariable.dicts('y',((i,j,k) for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres for k in Vehicle_Types if i!=j),lowBound=0)

    # Amount of delivery load across Arc(i,j) done by a Vehicle in Layer k
    z=p.LpVariable.dicts('z',((i,j,k) for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres for k in Vehicle_Types if i!=j),lowBound=0)

    # Set Objective Function (Point 2)
    prob+=p.lpSum(x[i,j,k]*C[i,j,k]*VS[k] for k in Vehicle_Types for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres if i!=j)+p.lpSum(x[0,j,k]*VC[k] for k in Vehicle_Types for j in Relief_Centres)

    #Ensuring at most a single vehicle caters to a Relief Center (Point 3 a)
    for i in Relief_Centres:
        prob+=p.lpSum(x[i,j,k] for j in Depot_and_Relief_Centres for k in Vehicle_Types if i!=j)<=1

    # Ensuring equal number of Incoming and Outgoing paths are available from all Nodes (Point 3 b)
    for i in Depot_and_Relief_Centres:
        for k in Vehicle_Types:
            prob+=p.lpSum(x[i,j,k]-x[j,i,k] for j in Depot_and_Relief_Centres  if i!=j)==0

    # Ensuring at most VN outgoing paths are available at the Depot since there are VN[k] vehicle for each Vehicle Type (Point 3 c)
    for k in Vehicle_Types:
        prob+=p.lpSum(x[0,j,k] for j in Relief_Centres)<=VN[k]

    """Flow Limitation Constraints"""

    for k in Vehicle_Types:
        for j in Relief_Centres:
            prob+=y[0,j,k]==0   #Ensuring initial PickUp from Nodes is 0 (Point 3 d i)

    for k in Vehicle_Types:
        for i in Relief_Centres:
            prob+=z[i,0,k]==0   #Ensuring final Delivery to Nodes is 0 (Point 3 d ii)

    #Ensuring the PickUp constraints are satisfied (Point 3 e i)
    for i in Relief_Centres:
        prob+=p.lpSum(y[i,j,k]-y[j,i,k] for j in Depot_and_Relief_Centres for k in Vehicle_Types if i!=j)==PickUp[i]

    #Ensuring the Delivery constraints are satisfied (Point 3 e ii)
    for i in Relief_Centres:
        prob+=p.lpSum(z[j,i,k]-z[i,j,k] for j in Depot_and_Relief_Centres for k in Vehicle_Types if i!=j)==Delivery[i]

    """Constraining the Sum of Flows to and from the Origin/Depot/Warehouse/NDRF_BASE"""
    #(Point 3 f i)
    prob+=p.lpSum(y[i,0,k] for i in Relief_Centres for k in Vehicle_Types)==p.lpSum(PickUp[i] for i in Relief_Centres)   #Ensuring sum of all PickUp Flow Variables to the Origin [0th Node] is equal to the total PickUps of all Nodes
    #(Point 3 f ii)
    prob+=p.lpSum(z[0,i,k] for i in Relief_Centres for k in Vehicle_Types)==p.lpSum(Delivery[i] for i in Relief_Centres)   #Ensuring sum of all Delivery Flow Variables from the Origin [0th Node] is equal to the total Demand of all Nodes

    # Ensuring the vehicle capacity is never exceeded (Point 3 g)
    for i in Depot_and_Relief_Centres:
        for j in Depot_and_Relief_Centres:
            for k in Vehicle_Types:
                if i!=j:
                    prob+=y[i,j,k]+z[i,j,k]<=VQ[k]*x[i,j,k]

    # Solve the Problem using default CBC
    #status=prob.solve(p.PULP_CBC_CMD(maxSeconds=max_seconds_allowed_for_calculation, msg=1, gapRel=0))
    if max_seconds_allowed_for_calculation>0:
        status=prob.solve(p.PULP_CBC_CMD(timeLimit=max_seconds_allowed_for_calculation))
    else:
        pulp_start_time=time.time()
        status=prob.solve()
        pulp_end_time=time.time()
    #v=len(Vehicle_Types)+len(Depot_and_Relief_Centres)
    #winsound.Beep(333+19*v, 777+11*v) # where 500 is the frequency in Hertz and 1000 is the duration in miliseconds
    print("This is the status:- ", p.LpStatus[prob.status])
    objec_val=p.value(prob.objective)


    Depot_First_Node=Depot_and_Relief_Centres[0]
    # Plotting the Depot and Relief Centres
    for k in Vehicle_Types:
        plt.figure(figsize=(11,11))
        for i in Depot_and_Relief_Centres:
            if i==Depot_First_Node:
                plt.scatter(Longitude[i],Latitude[i], c='r',marker='s')
                plt.text(Longitude[i] + 0.33, Latitude[i] + 0.33, "Depot")
            else:
                plt.scatter(Longitude[i], Latitude[i], c='black')
                plt.text(Longitude[i] + 0.33, Latitude[i] + 0.33, i)
        plt.title('mVRPSDC Tours for Vehicles of Type '+str(k)+" on the corresponding layer "+str(k))
        plt.ylabel("Latitude")
        plt.xlabel("Longitude")

        #routes = [(i, j) for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres  if i!=j if p.value(x[i,j,k])==1]
        routes = [(i, j) for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres  if i!=j and p.value(x[i,j,k])==1]

        
        # Drawing the optimal routes Layerwise
        Current_Node=Depot_First_Node
        colour_intervals=len(routes)
        for counter in range(colour_intervals):
            for i,j in routes:
                if i==Current_Node:
                    plt.annotate('', xy=[Longitude[j], Latitude[j]], xytext=[Longitude[i], Latitude[i]], arrowprops=dict(arrowstyle="-|>", connectionstyle='arc3', edgecolor=(counter/colour_intervals,1-(counter/colour_intervals),1)))
                    #plt.annotate('', xy=[Longitude[j], Latitude[j]], xytext=[Longitude[i], Latitude[i]], arrowprops=dict(arrowstyle="simple", connectionstyle='angle3', edgecolor=(counter/colour_intervals,1-(counter/colour_intervals),1)))
                    Edge_Notes="P="+str(y[i,j,k].varValue)+" ; D="+str(z[i,j,k].varValue)
                    plt.text((Longitude[i]+Longitude[j])/2, (Latitude[i]+Latitude[j])/2, f'{Edge_Notes}',fontweight="bold")
                    Current_Node=j
                    break
            routes.remove((i,j))

        max=0   # Finding the Maximum Utilised Vehicle Capacity for this Vehicle Type
        for i, j in routes:
            utilized_capacity=y[i,j,k].varValue+z[i,j,k].varValue
            if utilized_capacity>max:
                max=utilized_capacity
        Vehicle_Type_Maximum_Utilised_Capacity[k]=max
        print("The maximum vehicle capacity utilised ever in any tour in layer ",k," is: ",max," out of the total available",VQ[k])

        used_vehicles=0 # Finding the maximum number of vehicles being used
        for j in Relief_Centres:
            used_vehicles=p.value(x[0,j,k])+used_vehicles
        if used_vehicles!=0:
            Number_of_Vehicle_used_of_each_Type[k]=used_vehicles
        print("The maximum numbers of vehicles used is: ",used_vehicles," out of total available ",VN[k])
        if max_seconds_allowed_for_calculation>0:
            name="Vehicles_ "+str(used_vehicles)+"--"+str(VN[k])+" and Capacity_ "+str(max)+"--"+str(VQ[k])+".png"
        else:
            name="Vehicles_ "+str(used_vehicles)+"--"+str(VN[k])+" and Capacity_ "+str(max)+"--"+str(VQ[k])+".png"
        main_dir_for_Image=directory_to_save_PuLP_solution+"{}"
        plt.savefig(main_dir_for_Image.format(name))

    # Writing the Routes in a Text File

    textfile = open(directory_to_save_PuLP_solution+"Vehicle Routes as per PuLP.txt","w")
    if max_seconds_allowed_for_calculation>0:
        textfile.write("The Status of the problem is "+p.LpStatus[prob.status]+" which was stopped in time.\n")
        textfile.write("The Objective Value is "+str(objec_val)+" obtained within "+str(max_seconds_allowed_for_calculation)+" seconds with some Relative Gap \n")
    else:
        textfile.write("The Status of the problem is "+p.LpStatus[prob.status]+" \n")
        textfile.write("The Objective Value is "+str(objec_val)+" obtained within "+str(pulp_end_time-pulp_start_time)+" seconds \n")
    for k in Vehicle_Types:
        counter=0
        for j in Relief_Centres:
            if p.value(x[0,j,k])==1:
                counter+=1
                start_node=j
                textfile.write("Vehicle Type: "+str(k)+",\t Vehicle Number: "+str(counter)+", \t Route=\t 0")
                while start_node!=0:
                    textfile.write(" --> "+str(start_node))
                    for i in Depot_and_Relief_Centres:
                        if  start_node!=i and p.value(x[start_node,i,k])==1:
                            start_node=i
                            break
                if start_node==0:
                    textfile.write(" --> "+str(start_node)+"\n")
    for k in Vehicle_Type_Maximum_Utilised_Capacity:
        if Vehicle_Type_Maximum_Utilised_Capacity[k]>0:
            textfile.write("\n Vehicle Type "+str(k)+": Maximum Utilised Capacity by any Vehicle = "+Vehicle_Type_Maximum_Utilised_Capacity[k]+" units out of the total available "+VQ[k]+",\t and Number of Vehicles used is "+Number_of_Vehicle_used_of_each_Type[k]+" out of Total allowed "+VN[k])
            #textfile.write("\n Maximum Vehicle Capacity Utilised by any Vehicle of Type "+str(k)+" is "+Vehicle_Type_Maximum_Utilised_Capacity[k]+" units out of the total available "+VQ[k]+"\n")
    textfile.close()


    # Call a Workbook() function of openpyxl to create a new blank Workbook object
    wb_individual = openpyxl.Workbook()
    # Get workbook active sheet from the active attribute
    sheet_individual = wb_individual.active
    row_number_on_Individual_Sheet=1
    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 1)
    cell.value = "From Node i"
    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 2)
    cell.value = "To Node j"
    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 3)
    cell.value = "Vehicle Type k"
    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 4)
    cell.value = "x_ijk indicating whether the Arc is selected"
    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 5)
    cell.value = "y_ijk indicating the amount of Pickup"
    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 6)
    cell.value = "z_ijk indicating the amount of Delivery"
    for i in Depot_and_Relief_Centres:
        for j in Depot_and_Relief_Centres:
            for k in Vehicle_Types:
                if i!=j:
                    row_number_on_Individual_Sheet+=1
                    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 1)
                    cell.value = i
                    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 2)
                    cell.value = j
                    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 3)
                    cell.value = k
                    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 4)
                    cell.value = p.value(x[i,j,k])
                    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 5)
                    cell.value = p.value(y[i,j,k])
                    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 6)
                    cell.value = p.value(z[i,j,k])
    wb_individual.save(str(directory_to_save_PuLP_solution)+"Solution Details.xlsx")

    return objec_val,Vehicle_Type_Maximum_Utilised_Capacity,Number_of_Vehicle_used_of_each_Type