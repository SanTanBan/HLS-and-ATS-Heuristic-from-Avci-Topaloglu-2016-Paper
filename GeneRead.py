from openpyxl import Workbook
import pandas as pd
from random import random

# Indexes always start from 0 and are Integers only
# For the Distance Combination Reader, we need the Vehicle Type Set or it is obtained from the Excel File containing Vehicle Specifications which should be in the same folder
# All FILENAMES (whether Generated or Read) are within this single Python file which must be strictly followed accordingly

class Generator:
    def Vehicle_Type_Generator(array_of_Vehicle_Type_Indexes_considered=(),location_of_Original_Vehicle_Type_file="All Vehicle Type Specifications.xlsx",directory_to_save_Vehicle_Types_considered=""):
        wb=Workbook()
        ws=wb.create_sheet("Vehicle Specifications")
        output_from_All_Vehicle_Types_reader=Reader.Vehicle_Type_Reader(location_of_Vehicle_Type_file=location_of_Original_Vehicle_Type_file)
        row_number=1
        cell = ws.cell(row= row_number, column = 1)
        cell.value = "Vehicle Type"
        cell = ws.cell(row= row_number, column = 2)
        cell.value = "VN"
        cell = ws.cell(row= row_number, column = 3)
        cell.value = "VQ"
        cell = ws.cell(row= row_number, column = 4)
        cell.value = "VS"
        cell = ws.cell(row= row_number, column = 5)
        cell.value = "VC"
        cell = ws.cell(row= row_number, column = 6)
        cell.value = "Vehicle Width"
        for i in array_of_Vehicle_Type_Indexes_considered:
            row_number+=1
            cell = ws.cell(row= row_number, column = 1)
            cell.value = i
            cell = ws.cell(row= row_number, column = 2)
            cell.value = output_from_All_Vehicle_Types_reader[1][i]
            cell = ws.cell(row= row_number, column = 3)
            cell.value = output_from_All_Vehicle_Types_reader[2][i]
            cell = ws.cell(row= row_number, column = 4)
            cell.value = output_from_All_Vehicle_Types_reader[3][i]
            cell = ws.cell(row= row_number, column = 5)
            cell.value = output_from_All_Vehicle_Types_reader[4][i]
            cell = ws.cell(row= row_number, column = 6)
            cell.value = output_from_All_Vehicle_Types_reader[5][i]
        del wb["Sheet"]
        wb.save(str(directory_to_save_Vehicle_Types_considered)+"Vehicle Type Specifications.xlsx")

    def Node_Generator(upto_Node_Number,latitude_uniform_distribution_upper_bound=99999,latitude_uniform_distribution_lower_bound=0,pickup_quantity_uniform_distribution_lower_bound=0,pickup_quantity_uniform_distribution_upper_bound=99999,longitude_uniform_distribution_upper_bound=99999,longitude_uniform_distribution_lower_bound=0,delivery_quantity_uniform_distribution_lower_bound=0,delivery_quantity_uniform_distribution_upper_bound=99999,directory_location_to_be_saved=""):
        # Creating the Nodes for this instance
        wb=Workbook()
        ws=wb.create_sheet("Locations, PickUp & Delivery")
        # Defining the headings of the spreadsheet containing the Node details
        row_number=1
        cell = ws.cell(row= row_number, column = 1)
        cell.value = "Node Number"
        cell = ws.cell(row= row_number, column = 2)
        cell.value = "Latitude"
        cell = ws.cell(row= row_number, column = 3)
        cell.value = "Longitude"
        cell = ws.cell(row= row_number, column = 4)
        cell.value = "PickUp"
        cell = ws.cell(row= row_number, column = 5)
        cell.value = "Delivery"
        # Initialising the 0th Node as the Depot wihout any PickUp or Delivery Values
        row_number=2
        cell = ws.cell(row= row_number, column = 1)
        cell.value = "0"
        cell = ws.cell(row= row_number, column = 2)
        cell.value = str(latitude_uniform_distribution_lower_bound+random()*(latitude_uniform_distribution_upper_bound-latitude_uniform_distribution_lower_bound))
        cell = ws.cell(row= row_number, column = 3)
        cell.value = str(longitude_uniform_distribution_lower_bound+random()*(longitude_uniform_distribution_upper_bound-longitude_uniform_distribution_lower_bound))
        for i in range(1,upto_Node_Number+1):
            row_number+=1
            cell = ws.cell(row= row_number, column = 1)
            cell.value = str(i)
            cell = ws.cell(row= row_number, column = 2)
            cell.value = str(latitude_uniform_distribution_lower_bound+random()*(latitude_uniform_distribution_upper_bound-latitude_uniform_distribution_lower_bound))
            cell = ws.cell(row= row_number, column = 3)
            cell.value = str(longitude_uniform_distribution_lower_bound+random()*(longitude_uniform_distribution_upper_bound-longitude_uniform_distribution_lower_bound))
            cell = ws.cell(row= row_number, column = 4)
            cell.value = str(pickup_quantity_uniform_distribution_lower_bound+random()*(pickup_quantity_uniform_distribution_upper_bound-pickup_quantity_uniform_distribution_lower_bound))
            cell = ws.cell(row= row_number, column = 5)
            cell.value = str(delivery_quantity_uniform_distribution_lower_bound+random()*(delivery_quantity_uniform_distribution_upper_bound-delivery_quantity_uniform_distribution_lower_bound))
        del wb["Sheet"]
        wb.save(str(directory_location_to_be_saved)+"Node Specifications.xlsx")

    # Generating the Combinations as per the data available in the paper for all Vehicle Types
    def Euclidean_Distance_Matrix_Generator(location_of_Node_Locations_file="Node Specifications.xlsx",destination_to_save_Distance_Matrix=""):
        Relief_Locations=pd.read_excel(location_of_Node_Locations_file,"Locations, PickUp & Delivery",index_col=0)
        wb = Workbook()
        ws = wb.create_sheet("Euclidean Distances")
        row_number=1
        cell = ws.cell(row= row_number, column = 1)
        cell.value = "Origin Node"
        cell = ws.cell(row= row_number, column = 2)
        cell.value = "Destination Node"
        cell = ws.cell(row= row_number, column = 3)
        cell.value = "Distance"
        for i, row1 in Relief_Locations.iterrows():
            for j, row2 in Relief_Locations.iterrows():
                    if i!=j:
                        row_number+=1
                        cell = ws.cell(row= row_number, column = 1)
                        cell.value = str(i)
                        cell = ws.cell(row= row_number, column = 2)
                        cell.value = str(j)
                        cell = ws.cell(row= row_number, column = 3)
                        cell.value = ((row1["Latitude"]-row2["Latitude"])**2+(row1["Longitude"]-row2["Longitude"])**2)**0.5
        # This Distance Matrices are Euclidean Distances to be used by all Vehicle Types
        del wb["Sheet"]
        wb.save(str(destination_to_save_Distance_Matrix)+"Network Distance Matrix.xlsx")

    # Generating the Combinations as per the data available in the paper
    # here if p=0, then Random Distances are assigned in between p=1 and p=2, Otherwise the provided value of p ia used for computation of the Lp norm
    def Lp_Norm_Random_Matrix_for_each_Vehicle_Type_Generator(p=0,Vehicle_Types=[],location_of_Node_Locations_file="Node Specifications.xlsx",sheet_name_in_Node_Locations_file_containing_CoOrdinates="Locations, PickUp & Delivery",destination_to_save_Distance_Matrix=""):
        if p==0:
            pp=0
            print("Since p value is not mentioned, the returned distances (Lp norms) shall be between the Euclidean and Manhattan Distances")
        Relief_Locations=pd.read_excel(location_of_Node_Locations_file,sheet_name_in_Node_Locations_file_containing_CoOrdinates,index_col=0)
        wb = Workbook()
        for k in Vehicle_Types:
            ws = wb.create_sheet("Distances for VT "+str(k))
            row_number=1
            cell = ws.cell(row= row_number, column = 1)
            cell.value = "Origin Node"
            cell = ws.cell(row= row_number, column = 2)
            cell.value = "Destination Node"
            cell = ws.cell(row= row_number, column = 3)
            cell.value = "Distance"
            for i, row1 in Relief_Locations.iterrows():
                for j, row2 in Relief_Locations.iterrows():
                        if i!=j:
                            row_number+=1
                            cell = ws.cell(row= row_number, column = 1)
                            cell.value = str(i)
                            cell = ws.cell(row= row_number, column = 2)
                            cell.value = str(j)
                            cell = ws.cell(row= row_number, column = 3)
                            if pp==0:
                                p=random()+1
                            cell.value = ((row1["Latitude"]-row2["Latitude"])**p+(row1["Longitude"]-row2["Longitude"])**p)**(1/p)
            # This Distance Matrices can take random distances between Euclidean and Manhatan Distances which correspond to different network layers being used by different Vehicle Types
        del wb["Sheet"]
        wb.save(str(destination_to_save_Distance_Matrix)+"Network Distance Matrix.xlsx")

class Reader:
    #def Relief_Centre_Reader(location_of_Relief_Centre_Specifications_file="Node Specifications.xlsx",sheet_name_in_Relief_Centre_file_containing_CoOrdinates_PickUp_and_Delivery="Locations, PickUp & Delivery"):
    def Relief_Centre_Reader(directory_of_Relief_Centre_Specifications_file="",sheet_name_in_Relief_Centre_file_containing_CoOrdinates_PickUp_and_Delivery=[0]):
        # To be Updated to include Index of the Relief Centres, Location of each Relief Centre, Quantity of Each Product Needed, Left_Over_Capacity_to_Recieve_Human_Evacuees
        Nodes=pd.read_excel(directory_of_Relief_Centre_Specifications_file+"Node Specifications.xlsx",sheet_name_in_Relief_Centre_file_containing_CoOrdinates_PickUp_and_Delivery,index_col=0)
        Relief_Centres=[]
        Latitude={}
        Longitude={}
        PickUp={}
        Delivery={}
        for sheet in Nodes:
            for i,row in Nodes[sheet].iterrows():
                Relief_Centres.append(i)
                Latitude[i]=row["Latitude"]
                Longitude[i]=row["Longitude"]
                PickUp[i]=row["PickUp"]
                Delivery[i]=row["Delivery"]
        #return type 1 Array and 4 Dictionaries
        return Relief_Centres,Latitude,Longitude,PickUp,Delivery
        

    def Evacuation_Point_Reader(location_of_Evacuation_Point_Specifications_file=""):
        # Contains Index of the Evacuation Point, Location of each Evacuation Point, Quantity of Each Product Needed, Number of Evacuees to PickUp
        pass
    
    #def Vehicle_Type_Reader(location_of_Vehicle_Type_file="Vehicle Type Specifications.xlsx",sheet_name_in_Vehicle_Type_file="Vehicle Specifications"):
    def Vehicle_Type_Reader(directory_of_Vehicle_Type_file="",sheet_names_in_Vehicle_Type_file=[0]):
    #def Vehicle_Type_Reader(location_of_Vehicle_Type_file="Vehicle Type Specifications.xlsx"):
        # To be Updated to Contain Index of Vehicle Types, Name of Vehicle Type, Number of Vehicle Types Available at Each Vehicle Depot, Maximum Weight allowed into this Vehicle, Maximum Volume Allowed into this Vehicle, Variable Costs, Fixed Costs,  Maximum Speeds, Average Speeds, Variable Emissions, Fixed Emissions, Vehicle Width
        Vehicles=pd.read_excel(directory_of_Vehicle_Type_file+"Vehicle Type Specifications.xlsx",sheet_names_in_Vehicle_Type_file,index_col=0)
        Vehicle_Types=[]
        VN={}
        VQ={}
        VS={}
        VC={}
        Vehicle_Width={}
        for sheet in Vehicles:
            for i, row in Vehicles[sheet].iterrows():
                Vehicle_Types.append(i)
                VN[i]=row["VN"]
                VQ[i]=row["VQ"]
                VS[i]=row["VS"]
                VC[i]=row["VC"]
                Vehicle_Width=row["Vehicle Width"]
        # return type 1 Array containing the Indexes of the Vehicles, 5 Dictionaries
        return Vehicle_Types,VN,VQ,VS,VC,Vehicle_Width

    def Warehouse_Reader(location_of_Warehouse_Locations_file):
        # Contains Index of the Warehouse, Location of the Warehouse, Quantity of each Type of Product Available at the Warehouse
        pass
        
    def Vehicle_Depot_Reader(location_of_Vehicle_Depot_file):
        # Contains Index of the Vehicle Depots, Locations of the Vehicle Depots, Number of Vehicles of Each Type Available at the Depot
        pass
    
    def TransShipment_Point_Reader(location_of_TransShipment_Point_Location_file):
        # Contains Index of the Vehicle Depots, Locations of the Vehicle Depots, Number of Vehicles of Each Type Available at the Depot
        pass

    def Product_Specification_Reader(location_of_Product_Specification_file):
        # Contains Index of the Products, Weight of the Products, Volume of the Products, Priority of the Product (Clothes should have lesser Priority than Water)
        pass

    def Distance_Combinations_Reader(sheet_names=None,directory_containing_Distance_Locations_file="",directory_of_Vehicle_Types_considered_file=""):
    # sheet_name=None is to read all Sheets
        # Contains The Origin Node, Destination Node and the Distances of Paths
        # Assumed that the sequence of the Vehicle Type Indexes are followed in the Excel Sheet for Distance Matrix as well for each Vehicle Type to Distance Matrix creation
        # Distance Matrix we have created, i.e. for the C[i,j,k] will have k varying among the Vehicle Indexes
        output_of_Vehicle_Type_reader=Reader.Vehicle_Type_Reader(directory_of_Vehicle_Type_file=directory_of_Vehicle_Types_considered_file)
        Distances=pd.read_excel(directory_containing_Distance_Locations_file+"Network Distance Matrix.xlsx",sheet_names)
        k=0
        Distance_Matrix={}
        for sheet in Distances:
            for i,row in Distances[sheet].iterrows():
                Distance_Matrix[(row["Origin Node"],row["Destination Node"],output_of_Vehicle_Type_reader[0][k])]=row["Distance"]
            k+=1
        # returns 1 Dictionary containing the keys as a tuple of 
        return Distance_Matrix